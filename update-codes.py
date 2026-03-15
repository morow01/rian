#!/usr/bin/env python3
"""
Convert codes.xlsx → codes.json for the TimeSheet PWA.

Usage:
    python3 update-codes.py

Reads 'codes.xlsx' in the same folder and writes 'codes.json'.
The app loads codes.json on startup.

First run will auto-create a local .venv with openpyxl if needed.
"""
import sys, os, subprocess

script_dir = os.path.dirname(os.path.abspath(__file__))
venv_python = os.path.join(script_dir, '.venv', 'bin', 'python3')

# If openpyxl missing, bootstrap a local venv and re-run inside it
try:
    import openpyxl
except ImportError:
    if os.path.abspath(sys.executable) == os.path.abspath(venv_python):
        sys.exit("ERROR: openpyxl still not found in venv — something went wrong.")
    print("Setting up local .venv with openpyxl (one-time setup)...")
    subprocess.run([sys.executable, '-m', 'venv', os.path.join(script_dir, '.venv')], check=True)
    subprocess.run([venv_python, '-m', 'pip', 'install', 'openpyxl', '-q'], check=True)
    print("Done. Re-running script...\n")
    os.execv(venv_python, [venv_python] + sys.argv)

import json
from openpyxl import load_workbook

xlsx_path = os.path.join(script_dir, 'codes.xlsx')
json_path = os.path.join(script_dir, 'codes.json')

if not os.path.exists(xlsx_path):
    sys.exit(f"codes.xlsx not found at: {xlsx_path}")

wb = load_workbook(xlsx_path, read_only=True, data_only=True)
ws = wb['Work Codes']

rows = list(ws.iter_rows(min_row=2, values_only=True))
codes = []
skipped = 0

for i, row in enumerate(rows, 2):
    cat, desc, code = (str(v).strip() if v else '' for v in (row[0], row[1], row[2]))
    if not cat and not desc and not code:
        continue
    if not cat or not desc or not code:
        print(f"  WARNING row {i}: incomplete entry skipped — cat={repr(cat)} desc={repr(desc)} code={repr(code)}")
        skipped += 1
        continue
    codes.append({"category": cat, "description": desc, "code": code})

wb.close()

with open(json_path, 'w') as f:
    json.dump(codes, f, indent=2)

print(f"✓ Written {len(codes)} codes to codes.json" + (f" ({skipped} skipped)" if skipped else ""))
print(f"\nNext step — push to GitHub:")
print(f"  git add codes.json && git commit -m 'Update codes' && git push")
